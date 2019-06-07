# -*- coding: utf-8 -*-
"""8_lang_in_8_week.ipynb

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/1BasViB_tJiktvQde6_6Fm24kULgHM0mB

openpyxl -- >>   create a **workbook** *
"""

from openpyxl import Workbook
wb = Workbook()

""" ____________________create a worksheet________________________________"""

ws = wb.active   ##########  workbook is always created with at least one worksheet.
#---or----
ws1 = wb.create_sheet('my sheet name')
ws2 = wb.create_sheet('my second sheet name')
print(wb.sheetnames)

ws.title = 'new sheet'
print(wb.sheetnames)    # print active workbooks sheet names




# ----------------------------PLAYING WITH DATA ----------PLAYING WITH DATA----------------------------------------------



c = ws['A4']   # access the A4 cell , if not availble then create it
ws['A4'] = 4       # assigns the value
d = ws.cell(row=4, column=2, value=10)


#   important link --->> https://medium.com/aubergine-solutions/working-with-excel-sheets-in-python-using-openpyxl-4f9fd32de87f
#   important link ---->> https://www.geeksforgeeks.org/python-reading-excel-file-using-openpyxl-module/


for x in range(1,11):
        for y in range(1,15):
            ws.cell(row=x, column=y, value = x**2 + y**2)

        
# get max row count
max_row=ws.max_row
# get max column count
max_column=ws.max_column
# iterate over all cells 
# iterate over all rows
print('\n\n\n printing ws  \n\n')
for i in range(1,max_row+1):
     
     # iterate over all columns
     for j in range(1,max_column+1):
          # get particular cell value    
          cell_obj=ws.cell(row=i,column=j)
          # print cell value     
          print(cell_obj.value,end=' | ')
      #print new line
     print('\n')

"""----------SMS MAILER--------------"""

pip install twilio

# we import the Twilio client from the dependency we just installed
from twilio.rest import Client

# the following line needs your Twilio Account SID and Auth Token
client = Client("AC0########################", "2#########################cba")

# change the "from_" number to your Twilio number and the "to" number
# to the phone number you signed up for Twilio with, or upgrade your
# account to send SMS to any phone number
client.messages.create(to="+91-----------", 
                       from_="+12----------", 
                       body="Hello to myself from Python!")

"""Send SMS updates to mobile phone using python
If you are running any python script and want to send regular updates from your script to mobile phone through SMS, you can use SinchSMS API to send SMS.

Approach :
Create a app on Sinch and get the key and secret of the app and use these credentials in the following script to send SMS to your mobile.

Limitation of Sinch :
If you don’t have any credits(you have to pay for credits), you can only send SMS to the registered mobile numbers on Sinch.
You can use way2sms to send sms to any number(I will be discussing how to use way2sms in another article), but without purchased credits, on way2sms also, you can’t send more than 100 SMS per day
"""

# another way to send message( nearly same process as above)
# python script for sending message update 
# 

import time 
from time import sleep 
from sinchsms import SinchSMS 

# function for sending SMS 
def sendSMS(): 

	# enter all the details 
	# get app_key and app_secret by registering 
	# a app on sinchSMS 
	number = 'your_mobile_number'
	app_key = 'your_app_key'
	app_secret = 'your_app_secret'

	# enter the message to be sent 
	message = 'Hello Message!!!'

	client = SinchSMS(app_key, app_secret) 
	print("Sending '%s' to %s" % (message, number)) 

	response = client.send_message(number, message) 
	message_id = response['messageId'] 
	response = client.check_status(message_id) 

	# keep trying unless the status retured is Successful 
	while response['status'] != 'Successful': 
		print(response['status']) 
		time.sleep(1) 
		response = client.check_status(message_id) 

	print(response['status']) 

if __name__ == "__main__": 
	sendSMS()

### making confusion matrix without importing any library.

import numpy as np

true_data = np.array([1,0,0,1,0,0,1,1,1,0],int)
machine_data=np.array([1,1,0,1,0,0,1,0,0,0],int)
def cm(true_data,machine_data):
        l = len(true_data)
        a=0;  b=0;   c=0;   d=0
        for i in range(l):
            if true_data[i]==1 and machine_data[i] ==1:
                a+=1
            if true_data[i]==1 and machine_data[i] ==0:
                b+=1
            if true_data[i]==0 and machine_data[i] ==1:
                c+=1
            if true_data[i]==0 and machine_data[i] ==0:
                d+=1
        result = np.array([[a,b],[c,d]])
        return result
x = cm(true_data,machine_data)
print(x)

## confusion matrix from sklearn

from sklearn.metrics import confusion_matrix
import numpy as np
true_data = np.array([1,0,0,1,0,0,1,1,1,0],int)
machine_data=np.array([1,1,0,1,0,0,1,0,0,0],int)
result = confusion_matrix(true_data,machine_data)
print(result)

# -------------------TIC TAC TOE ----------

from __future__ import print_function

choices = []

for x in range (0, 9) :
    choices.append(str(x + 1))

playerOneTurn = True
winner = False

def printBoard() :
    print( '\n --------')
    print( '|' + choices[0] + '|' + choices[1] + '|' + choices[2] + '|')
    print( ' -----')
    print( '|' + choices[3] + '|' + choices[4] + '|' + choices[5] + '|')
    print( ' -----')
    print( '|' + choices[6] + '|' + choices[7] + '|' + choices[8] + '|')
    print( ' --------\n')

while not winner :
    printBoard()

    if playerOneTurn :
        print( "Player 1:")
    else :
        print( "Player 2:")

    try:
        choice = int(input(">> "))
    except:
        print("please enter a valid field")
        continue
    if choices[choice - 1] == 'X' or choices [choice-1] == 'O':
        print("illegal move, plase try again")
        continue

    if playerOneTurn :
        choices[choice - 1] = 'X'
    else :
        choices[choice - 1] = 'O'

    playerOneTurn = not playerOneTurn

    for x in range (0,3):
        y = x * 3
        if choices[y]==choices[(y+1)] and choices[y]==choices[(y+2)]:
            winner = True
            printBoard()
        if (choices[x]==choices[(x+3)] and choices[x]==choices[(x+6)]) :
            winner = True
            printBoard()

    if((choices[0] == choices[4] and choices[0] == choices[8]) or 
       (choices[2] == choices[4] and choices[4] == choices[6])) :
        winner = True
        printBoard()

print ("Player " + str(int(playerOneTurn + 1)) + " wins!\n")
